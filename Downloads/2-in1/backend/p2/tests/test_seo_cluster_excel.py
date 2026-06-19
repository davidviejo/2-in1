import io

import pandas as pd
from openpyxl import load_workbook

from apps.web.blueprints.seo_tool import (
    FINAL_STRATEGY_SHEET_NAME,
    SERP_HISTORY_SHEET_NAME,
    build_final_strategy_dataframe,
    build_serp_history_dataframe,
    list_to_numbered_text,
    load_history,
    numbered_text_to_list,
    write_final_strategy_excel,
)


def _xlsx(sheets):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
    buf.seek(0)
    return buf


def test_numbered_text_helpers_dedupe_and_parse_google_sheets_paste():
    assert list_to_numbered_text([' kw uno ', '', 'kw dos', 'KW uno']) == '1. kw uno\n2. kw dos'
    assert numbered_text_to_list('1. kw uno 2. kw dos\r\n3) kw tres\n- kw dos') == ['kw uno', 'kw dos', 'kw tres']
    assert numbered_text_to_list('uno, dos; tres') == ['uno', 'dos', 'tres']


def test_import_new_single_sheet_parses_lists_and_dedupes_urls():
    wb = _xlsx({
        FINAL_STRATEGY_SHEET_NAME: pd.DataFrame([{
            'Cluster ID': 'G-001',
            'Keyword principal': 'keyword padre',
            'Variaciones': '1. keyword hija\n2. keyword hija\n3. keyword larga',
            'Nº keywords': 3,
            'Intención': 'Informacional',
            'Cobertura': 'Oportunidad',
            'URLs propias': '1. https://example.com/a\n2. https://example.com/a',
            'URLs SERP principales': '1. https://competidor.com/a\n2. https://competidor.com/b',
            'Nº URLs SERP': 2,
            'Títulos SERP principales': '1. Título A\n2. Título B',
            'Avg Palabras': 1000,
            'Avg Imágenes': 4,
            'Entidades': '1. Madrid\n2. SEO',
            'Estructura': '1. H1\n2. H2',
        }])
    })
    clusters, max_id = load_history(wb)
    assert max_id == 1
    assert clusters[0]['children'] == ['keyword hija', 'keyword larga']
    assert clusters[0]['own_urls'] == ['https://example.com/a']
    assert [u['url'] for u in clusters[0]['serp_dump']] == ['https://competidor.com/a', 'https://competidor.com/b']
    assert clusters[0]['entities'] == ['Madrid', 'SEO']


def test_import_legacy_multi_sheet_still_supported():
    wb = _xlsx({
        'Estrategia': pd.DataFrame([
            {'Cluster ID': 'G-002', 'Rol': 'PADRE', 'Keyword': 'padre', 'Avg Palabras': '-', 'Avg Imágenes': '-', 'Estructura': '-', 'Entidades': '-'},
            {'Cluster ID': 'G-002', 'Rol': 'Variación', 'Keyword': 'hija', 'Avg Palabras': '-', 'Avg Imágenes': '-', 'Estructura': '-', 'Entidades': '-'},
        ]),
        'URLs': pd.DataFrame([{'Cluster ID': 'G-002', 'Padre': 'padre', 'Rank': 1, 'URL': 'https://serp.test/a', 'Título': 'A'}]),
        'Overlap SERP': pd.DataFrame([{'Cluster ID': 'G-002', 'Keyword Padre': 'padre', 'Keyword Variación': 'hija', 'URLs Coincidentes': '1. https://serp.test/a'}]),
    })
    clusters, max_id = load_history(wb)
    assert max_id == 2
    assert clusters[0]['parent'] == 'padre'
    assert clusters[0]['children'] == ['hija']
    assert clusters[0]['serp_dump'][0]['url'] == 'https://serp.test/a'


def test_export_new_format_single_sheet_no_internal_columns_and_roundtrip():
    clusters = [{
        'id': 'G-010',
        'parent': 'padre',
        'children': ['hija', 'otra hija'],
        'intent': 'Informacional',
        'coverage': 'Cubierto',
        'own_urls': ['https://example.com/a', 'https://example.com/a'],
        'serp_dump': [{'url': 'https://serp.test/a', 'title': 'A', 'rank': 1}],
        'keyword_serps': {
            'padre': [{'url': 'https://serp.test/a', 'title': 'A', 'rank': 1}],
            'hija': [{'url': 'https://serp.test/b', 'title': 'B', 'rank': 1}],
        },
        'analyzed': True,
        'avg_words': 1200,
        'avg_imgs': 3,
        'entities': ['SEO', 'Contenido'],
        'top_structure': ['H1', 'H2'],
    }]
    df = build_final_strategy_dataframe(clusters)
    assert list(df.columns) == [
        'Cluster ID', 'Keyword principal', 'Variaciones', 'Nº keywords', 'Intención', 'Cobertura',
        'URLs propias', 'URLs SERP principales', 'Nº URLs SERP', 'Títulos SERP principales',
        'Avg Palabras', 'Avg Imágenes', 'Entidades', 'Estructura'
    ]
    history_df = build_serp_history_dataframe(clusters)
    assert history_df[['Keyword', 'Tipo keyword', 'URL']].to_dict('records') == [
        {'Keyword': 'padre', 'Tipo keyword': 'Principal', 'URL': 'https://serp.test/a'},
        {'Keyword': 'hija', 'Tipo keyword': 'Variación', 'URL': 'https://serp.test/b'},
    ]
    buf = io.BytesIO()
    write_final_strategy_excel(clusters, buf)
    buf.seek(0)
    wb = load_workbook(buf)
    assert wb.sheetnames == [FINAL_STRATEGY_SHEET_NAME, SERP_HISTORY_SHEET_NAME]
    ws = wb[FINAL_STRATEGY_SHEET_NAME]
    assert ws.freeze_panes == 'A2'
    assert ws.auto_filter.ref is not None
    assert ws['C2'].alignment.wrap_text is True
    history_ws = wb[SERP_HISTORY_SHEET_NAME]
    assert history_ws['B2'].value == 'padre'
    assert history_ws['C3'].value == 'Variación'
    assert history_ws['E3'].value == 'https://serp.test/b'
    buf.seek(0)
    imported, _ = load_history(buf)
    assert imported[0]['children'] == ['hija', 'otra hija']
    assert imported[0]['own_urls'] == ['https://example.com/a']
